# -*- coding: utf-8 -*-

from openpyxl import Workbook
import re
from datetime import datetime
import fileinput


index = 1
count = 0
t1 = datetime.now()
with open('C:\\Users\\e485858\\Downloads\\_PRODUCT.20200918.xml.20200919100049',mode='r',buffering=50000,errors='ignore') as file:
    for line in file:
        if index == 1:
            wb = Workbook()
            sheet1 = wb.create_sheet("Page0")
            
        string = line
        if string =='':
            break
        new_str = re.sub('[^a-zA-Z0-9\n\.\^\*\-]', ' ', string)
        arr = list(new_str.split('^'))
    
        for i in range(len(arr)):
            sheet1.cell(row=index,column=i+1).value =arr[i]
        
        index += 1
        if index == 100000:
            wbname = 'newwb'+str(count)+'.xlsx'
            wb.save('C:\\Users\\e485858\\Documents\\Work\\Python\\Workspace\\AssignmentProjects\\HackerRank\\Files\\'+wbname)
            index = 1
            count+=1
            print(count)

print('time lapse:',datetime.now() - t1)

